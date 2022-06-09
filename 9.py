# !/usr/bin/env python
# -*- coding:utf-8 -*-
# @Author  : Michael jok
# @file       :  1.py
# @Time    : 2022/4/16 10:11
# @Function:
# !/usr/bin/env python
# -*- coding:utf-8 -*-
# @Author  : Michael jok
# @file       :  1.py
# @Time    : 2022/2/28 18:12
# @Function:
import xlrd3
import xml.etree.ElementTree
from openpyxl.styles import PatternFill,Alignment
import os
import time
#获取当前文件夹路径

def jiangeleibi(jiange,lis1):
    # lisjiage = [lis1[0],lis1[1]-lis1[0],lis1[2]-lis1[1],lis1[3]-lis1[2],lis1[4]-lis1[3],lis1[5]-lis1[4]]
    lisjiage = [lis1[0],lis1[1]-lis1[0],lis1[2]-lis1[1],lis1[3]-lis1[2],]
    for biz,t in enumerate(jiange):
        if t<lisjiage[biz]:
            jiange[biz]=lisjiage[biz]
    return jiange


#基本情况总list:
jibenlist = []

fill = PatternFill(fill_type='solid', fgColor="d3d3d3", bgColor="1874CD")


paths = os.getcwd()

lists = os.listdir(f'{paths}/xlsx')


liss = ['个人职业发展目标','年度能力提升目标','计划承诺','over']

pages = {}

for ssf in liss:
    pages[ssf]=0
zongjiange = [0, 0, 0, 0]
for bbb in lists:
    zq = []
    if bbb.endswith(".xlsx"):
        data = xlrd3.open_workbook(f"{paths}/xlsx/{bbb}")  # 实例化
        sheet = data.sheet_by_index(0)
        hang = sheet.nrows
        liss = ['个人职业发展目标', '年度能力提升目标', '计划承诺', 'over']
        jianges = []
        for listid, i in enumerate(range(hang)):
            ppp = sheet.row_values(i)
            #循环页码
            for sidnf in liss:
                print(sidnf,ppp[0])
                if sidnf in ppp[0]:
                    jianges.append(listid)
        jianges.append(listid)
        print(jianges)
        print("啊哈哈",zongjiange,jianges)
        zongjiange = jiangeleibi(zongjiange,jianges)
print(zongjiange)

liuliu = 0
for sdfss,bbb in enumerate(pages):
    if sdfss==0:
        pages[bbb] = zongjiange[sdfss]
        liuliu = zongjiange[sdfss]
    else:
        pages[bbb] = liuliu+zongjiange[sdfss]
        liuliu = liuliu+zongjiange[sdfss]
pages['over']=pages['计划承诺']+2*(pages['over']-pages['计划承诺'])
#基本信息
lists1 = ['计划信息及承诺','人员信息','计划名称：','学习期限：','学员：','部门：','岗位：','导师：','']
print('list1',lists1)

#个人职业发展目标
def rs_lists3(zhiyefazhan,nenglitisheng):
    lists3 = ['总结您的三项优势项','总结您的三项不足项','三年左右职业发展目标']
    shu = (nenglitisheng - zhiyefazhan) // 3
    lists3 = lists3 * shu
    lists3.append("")
    lists3.insert(0, "个人职业发展目标")
    return lists3
lists3 = rs_lists3(pages['个人职业发展目标'],pages['年度能力提升目标'])
print('list3',lists3)


#能力提升学习发展目标（建议3~5项）
def rs_lists4(zhiyefazhan,nenglitisheng):
    lists3 = ['能力提升项（结合个人优劣势现状）','当前能力现状','能力提升具体目标（非业绩目标）',' ']
    shu = (nenglitisheng - zhiyefazhan) // 3
    lists3 = lists3 * shu
    lists3.append("")
    lists3.insert(0, "年度能力提升目标（点击新增添加目标项）")
    return lists3
lists4 = rs_lists4(pages['年度能力提升目标'],pages['计划承诺'])
print('list4',lists4)



#计划承诺
def rs_lists6(zhiyefazhan,nenglitisheng):
    lists3=['']
    shu = (nenglitisheng - zhiyefazhan)
    lists3 = lists3 * shu
    lists3.insert(0, "计划承诺")
    return lists3
lists6 = rs_lists6(pages['计划承诺'],pages['over'])
print('list6',lists6)

zonglist = lists1+lists3+lists4+lists6


print(zonglist)

# zonglist.remove('计划信息及承诺')
import openpyxl
wb = openpyxl.Workbook()
ppppp = 0

ws = wb.create_sheet("计划信息及承诺", ppppp)
ws.append(zonglist)
ws1zonghang = 1




lists =os.listdir(f'{paths}/xlsx')

for bbb in lists:
    zq = []
    if bbb.endswith(".xlsx"):
        data = xlrd3.open_workbook(f"{paths}/xlsx/{bbb}")  # 实例化
        sheet = data.sheet_by_index(0)
        hang = sheet.nrows
        #增加一列：
        douding = []
        for zhongsin in zonglist:
            douding.append('')

        lissa = ['个人职业发展目标', '年度能力提升目标', '计划承诺', 'over']
        pagesa = {}

        for ssfa in lissa:
            pagesa[ssfa] = 0


        for listid, i in enumerate(range(hang)):
            ppp = sheet.row_values(i)
            #循环页码
            for sdinf in pages:
                if sdinf in ppp[0]:
                    if listid>pagesa[sdinf]:
                        pagesa[sdinf] = listid
            if pagesa['over'] < listid:
                pagesa['over']=listid
        print(bbb)

        #获取基本信息
        t= 0
        jiben = []
        while t<pagesa['个人职业发展目标']:
            jiben.append(sheet.row_values(t)[1])
            t+=1
        print("基本情况",jiben)
        jibenlist.append(jiben[2:8])


        for qqq,bins in enumerate(jiben):
            print(qqq)
            douding[qqq] = bins




        # 获取个人现状
        t = pagesa['个人职业发展目标']
        zhiyefazhan = []
        while t < pagesa['年度能力提升目标'] and t >= pagesa['个人职业发展目标']:
            zhiyefazhan.append(sheet.row_values(t)[1])
            t += 1
        print("职业发展》》》",zhiyefazhan)

        tsss = pages['个人职业发展目标']
        for qqq, bins in enumerate(zhiyefazhan):
            douding[tsss + qqq] = bins
        # 能力提升学习发展目标
        t = pagesa['年度能力提升目标']
        nenglitisheng = []
        while t < pagesa['计划承诺'] and t >= pagesa['年度能力提升目标']:
            nenglitisheng.append(sheet.row_values(t)[1])
            t += 1
        print(nenglitisheng)
        print(len(nenglitisheng))

        tsss = pages['年度能力提升目标']
        print(tsss)
        for qqq, bins in enumerate(nenglitisheng):
            douding[tsss + qqq] = bins

        # 计划承诺
        t = pagesa['计划承诺']
        jihua = []
        lsdji = 0
        while t <= pagesa['over'] and t >= pagesa['计划承诺']:
            if lsdji == 0:
                jihua.append(' ')
                jihua.append(' ')
                lsdji += 1
            else:
                jihua.append(sheet.row_values(t)[0])
                jihua.append(sheet.row_values(t)[1])
                # jihua.append(' '.join(sheet.row_values(t)))
            t += 1
        print(jihua)
        # exit()
        tsss = pages['计划承诺']
        for qqq, bins in enumerate(jihua):
            douding[tsss + qqq] = bins
        ws.append(douding)
        ws1zonghang += 1

#判断是否为空列表
def listTF(lists):
    for b in lists:
        if str(b).strip():
            return True


# # for sdfss,oasdjif in enumerate(zhuanheng):
#     ws.append(oasdjif)

ws1 = wb.create_sheet("具体行动计划", 1)
# yanse = []
ws1list = []
tou = ['计划名称：', '学习期限：', '学员：', '部门：', '岗位：', '导师：','完成度：','任务名称','任务类型','任务详情','阶段名称','完成时间','学习心得','求助记录']
ws1list.append(tou)
for sdfef,bbb in enumerate(lists):
    zq = []
    if bbb.endswith(".xlsx"):
        data = xlrd3.open_workbook(f"{paths}/xlsx/{bbb}")  # 实例化

        sheet1 = data.sheet_by_index(1)
        jibenqingkuang = jibenlist.copy()[sdfef]
        hang = sheet1.nrows
        # print('一共',hang)

        for listid, i in enumerate(range(hang)):
            if listid==0:
                jibenqingkuang.append(sheet1.row_values(i)[4])
            if listid>1:
                beis = jibenqingkuang.copy()+sheet1.row_values(i)
                ws1list.append(beis)








yanse1 = []

for ccts,bbc in enumerate(ws1list):
    ws1.append(bbc)
    if '人员信息'in bbc[0]:
        yanse1.append(ccts+1)

for yanse_ in yanse1:
    ws1.cell(row=yanse_, column=1).fill = fill
    ws1.cell(row=yanse_, column=3).fill = fill
    ws1.cell(row=yanse_, column=6).fill = fill
#
# exit()

ws2 = wb.create_sheet("求助记录", 2)
ws2list = []
tou = ['计划名称：', '学习期限：', '学员：', '部门：', '岗位：', '导师：','任务名称','求助问题','问题描述','导师回复','是否解决']
ws2list.append(tou)
for sdfef,bbb in enumerate(lists):
    zq = []
    if bbb.endswith(".xlsx"):
        data = xlrd3.open_workbook(f"{paths}/xlsx/{bbb}")  # 实例化
        sheet2 = data.sheet_by_index(2)
        jibenqingkuang = jibenlist.copy()[sdfef]
        hang = sheet2.nrows
        for listid, i in enumerate(range(hang)):
            if listid > 1:
                beis = jibenqingkuang.copy() + sheet2.row_values(i)
                ws2list.append(beis)




yanse2 = []

for ccts, bbc in enumerate(ws2list):
    ws2.append(bbc)
    if '人员信息' in bbc[0]:
        yanse2.append(ccts + 1)

for yanse_ in yanse2:
    ws2.cell(row=yanse_, column=1).fill = fill
    ws2.cell(row=yanse_, column=3).fill = fill




ws3 = wb.create_sheet("计划及各阶段回顾", 3)
ws3list = []
tou = ['计划名称：', '学习期限：', '学员：', '部门：', '岗位：', '导师：','阶段','重点发展内容','达到的成果或效果','完成度打分','目标达成/未达成原因分析','导师点评']
ws3list.append(tou)
for sdfef,bbb in enumerate(lists):
    zq = []
    if bbb.endswith(".xlsx"):
        data = xlrd3.open_workbook(f"{paths}/xlsx/{bbb}")  # 实例化
        sheet3 = data.sheet_by_index(3)
        jibenqingkuang = jibenlist.copy()[sdfef]
        hang = sheet3.nrows
        for listid, i in enumerate(range(hang)):
            if listid > 1:
                beis = jibenqingkuang.copy() + sheet3.row_values(i)
                ws3list.append(beis)

yanse3 = []

for ccts, bbc in enumerate(ws3list):
    ws3.append(bbc)
    if '人员信息' in bbc[0]:
        yanse3.append(ccts + 1)

for yanse_ in yanse3:
    ws3.cell(row=yanse_, column=1).fill = fill
    ws3.cell(row=yanse_, column=3).fill = fill

ws4 = wb.create_sheet("学习心得", 4)
ws4list = []
tou = ['计划名称：', '学习期限：', '学员：', '部门：', '岗位：', '导师：','任务名称','心得内容','导师点评']
ws4list.append(tou)
for sdfef,bbb in enumerate(lists):
    zq = []
    if bbb.endswith(".xlsx"):
        data = xlrd3.open_workbook(f"{paths}/xlsx/{bbb}")  # 实例化
        sheet4 = data.sheet_by_index(3)
        jibenqingkuang = jibenlist.copy()[sdfef]
        hang = sheet4.nrows
        for listid, i in enumerate(range(hang)):
            if listid > 1:
                beis = jibenqingkuang.copy() + sheet4.row_values(i)
                ws4list.append(beis)
yanse4 = []

for ccts, bbc in enumerate(ws4list):
    ws4.append(bbc)
    if '人员信息' in bbc[0]:
        yanse4.append(ccts + 1)

for yanse_ in yanse4:
    ws4.cell(row=yanse_, column=1).fill = fill
    ws4.cell(row=yanse_, column=3).fill = fill


ws.cell(row=1, column=1).fill = fill
ws.cell(row=1, column=2).fill = fill
for bsidnf in liss[:-1]:
    if bsidnf!='计划承诺':
        shu = pages[bsidnf]+1
    else:
        shu=pages[bsidnf]+2
    ws.cell(row=1, column=shu).fill = fill
#第一个表是否自动换行
# for sdef,kkl in enumerate(zonglist):
#     for tt in range(2,ws1zonghang):
#         ws.cell(row=tt, column=(sdef+1)).alignment = Alignment(wrapText=True)


wb.save("oks.xlsx")
# print(jibenlist)
print("操作完毕，30秒后退出！")
time.sleep(30)


# print(pages)